import math
# import requests
# from django.http import HttpResponseRedirect
from django.shortcuts import render
from openpyxl import load_workbook
from .forms import PostForm
from .letter import cost_of_simple, cost_of_registered, cost_of_value_letter
from .first_class import cost_of_first_class


def read_letter_from_exel(filepath):
    price_list = []
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=3, values_only=True):
        item, cost_fiz, cost_yur = row
        rates = {
            'item': item,
            'cost_fiz': cost_fiz,
            'cost_yur': cost_yur
                    }
        price_list.append(rates)
    return price_list



def calculation_view(request):
    if request.method == "POST":
        form = PostForm(request.POST)
        if form.is_valid():
            item_weight = int(request.POST.get('weight'))
            declared_value = request.POST.get('declared_value')
            simple = cost_of_simple(item_weight)
            registered = cost_of_registered(item_weight)
            value_letter = cost_of_value_letter(item_weight, declared_value)
            first_class = cost_of_first_class(item_weight)
            context = {'form': form, 'simple': simple,
                       'registered': registered,
                       'value_letter': value_letter,
                       'first_class': first_class}
            print(context)
            return render(request, 'index.html', context)  # Внутри фиг скобок
    else:
        form = PostForm()
        return render(request, 'index.html', {'form': form})  # внутри фигурных скобок
