import os
from openpyxl import load_workbook


file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files/ems_points.xlsx')
wb = load_workbook(filename=file_path)
sheet = wb.active


def read_ems_points(worksheet):
    rows = sheet.max_row
    choices = []
    for i in range(1, rows + 1):
        cell = worksheet.cell(row=i, column=2)
        row = (i, cell.value)
        choices.append(row)
    choices = tuple(choices)
    return choices


def find_ems_zone(worksheet, point):
    # zone_list = []
    ems_point = 0
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        # zone_list.append(row)
        if str(row[0]) == point:
            ems_point = row[2]
    return ems_point


def cost_of_ems(departure, destination, item_weight, declared_value):
    zone1 = find_ems_zone(sheet, departure)
    zone2 = find_ems_zone(sheet, destination)
    return departure, destination, zone1, zone2, item_weight, declared_value


print(read_ems_points(sheet))
print(find_ems_zone(sheet, 5))
print(cost_of_ems(1, 5, 200, 10))
"""
правильно определяет зоны
"""
