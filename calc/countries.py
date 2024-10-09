import os
from openpyxl import load_workbook


file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files/parcel_int.xlsx')
wb = load_workbook(filename=file_path)
sheet2 = wb.active


def read_the_country(worksheet):
    rows = sheet2.max_row
    choices = []
    for i in range(11, rows + 1):
        cell = worksheet.cell(row=i, column=3)
        row = (i, cell.value)
        choices.append(row)
    choices = tuple(choices)
    return choices
