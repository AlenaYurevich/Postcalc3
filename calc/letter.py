import os
import math
from openpyxl import load_workbook
from .vat import vat
from .format import formatted
from .ems_cost import notification_list
from .declared_value import cost_for_declared_value


def weight_step(weight):
    return math.ceil((weight - 20) / 20)


def notification_cost(notification):
    print("обратились к функции notification_cost")
    match notification:
        case 1: return notification_list[0]
        case 2: return notification_list[1]
        case 3: return notification_list[2]
        case 4: return 0


def cost_of_simple(item_weight):
    price_row = []
    if item_weight <= 2000:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files/letter.xlsx')  # первый файл
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        fiz = sheet['B5'].value + sheet['B6'].value * weight_step(item_weight)
        yur = sheet['C5'].value + sheet['C6'].value * weight_step(item_weight)
        item_vat = vat(yur)
        yur += item_vat
        rate = {
            'fiz': fiz,
            'yur': yur,
            'item_vat': item_vat,
            'for_declared': "",
            'rub': " руб.",
            'tracking': "нет"
        }
        for i in rate:
            rate[i] = formatted(rate[i])
        price_row.append(rate)
    else:
        fiz = "Макс. вес 2 кг"
        price_row.append({'fiz': fiz})
    return price_row


"""
заказное письмо, заказная бандероль, заказной мелкий пакет
"""


def cost_of_registered(item_weight, notification):
    price_row = []
    notification = notification_cost(notification)
    if item_weight <= 2000:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files/letter2.xlsx')  # второй файл
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        fiz = sheet['D10'].value + sheet['D12'].value * weight_step(item_weight)
        yur = sheet['H10'].value + sheet['H12'].value * weight_step(item_weight)
        fiz += notification * 1.2
        yur += notification
        notification = notification * 1.2
        if notification == 0:
            notification = ""
        item_vat = vat(yur)
        yur += item_vat
        rate = {
            'fiz': fiz,
            'yur': yur,
            'item_vat': item_vat,
            'for_declared': "",
            'rub': " руб.",
            'tracking': "да",
            'notification': notification
        }
        for i in rate:
            rate[i] = formatted(rate[i])
        price_row.append(rate)
    else:
        fiz = "Макс. вес 2 кг"
        price_row.append({'fiz': fiz})
    return price_row


"""
Письмо, мелкий пакет с объявленной ценностью
проверить, можно ли перенести в новый файл
"""


def cost_of_value_letter(item_weight, declared_value, notification):
    price_row = []
    notification = notification_cost(notification)
    if item_weight <= 2000:
        if declared_value not in ("нет", "", 0, "0"):
            file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files/letter2.xlsx')  # второй файл
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            fiz = sheet['D11'].value + sheet['D12'].value * weight_step(item_weight)\
                                     + cost_for_declared_value(declared_value) * 1.2
            yur = sheet['H11'].value + sheet['H12'].value * weight_step(item_weight)\
                                     + cost_for_declared_value(declared_value)
            fiz += notification
            yur += notification
            notification = notification * 1.2
            if notification == 0:
                notification = ""
            item_vat = vat(yur)
            yur += item_vat
            for_declared = cost_for_declared_value(declared_value) * 1.2
            rate = {
                'fiz': fiz,
                'yur': yur,
                'item_vat': item_vat,
                'for_declared': for_declared,
                'rub': " руб.",
                'tracking': "да",
                'sep': '/',
                'notification': notification
            }
            for i in rate:
                rate[i] = formatted(rate[i])
            price_row.append(rate)
        else:
            rate = {
                'fiz': "",
                'yur': "",
                'item_vat': "",
                'for_declared': "",
                'rub': "",
                'sep': "",
                'notification': ""
            }
            price_row.append(rate)
    else:
        fiz = "Макс. вес 2 кг"
        price_row.append({'fiz': fiz, 'sep': ''})
    return price_row
