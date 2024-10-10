import os
import math
from openpyxl import load_workbook
from .vat import vat
from .round_as_excel import round_as_excel
from .format import formatted
from .declared_value import cost_for_declared_value


file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files/parcel_int.xlsx')
wb = load_workbook(filename=file_path)
sheet = wb.active


"""
для посылок стоимость складывается из стоимости за посылку и стоимости за килограмм
Тарификация  за  массу  посылки  без  объявленной  ценности  осуществляется  с точностью до сотен 
граммов. Любое количество граммов округляется до сотни граммов в большую сторону
Тарификация  за  массу  посылки с  объявленной  ценностью  осуществляется  с точностью  до  
десятков  граммов.  Любое  количество  граммов  округляется  до  десятков  граммов  в
большую сторону
"""


def weight(item_weight, declared_value):
    if declared_value in ("нет", "", 0, "0"):
        return math.ceil(item_weight / 100) / 10
    else:
        return math.ceil(item_weight / 10) / 100  # повторяется код с parcel


def find_numbers_by_country(row_number):
    if row_number < 11:
        # Если номер строки меньше 11, возвращаем None, потому что строки выше 11 не обрабатываются.
        return None, None
        # Получаем строку по её индексу.
    row_data = list(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True))

    if row_data:  # Проверка, что строка не пустая
        row_data = row_data[0]  # Извлекаем данные строки (поскольку iter_rows возвращает список кортежей)
        non_priority = (row_data[3], row_data[4])  # Получаем данные из 3-й и 4-й колонок
        priority = (row_data[5], row_data[6])  # Получаем данные из 5-й и 6-й колонок
        return non_priority, priority
    # Если row_data пустой или None, возвращаем None, None
    return None, None


def cost_of_parcel_int(destination, item_weight, declared_value):
    price_row = []
    if item_weight <= 50000:
        non_priority, priority = find_numbers_by_country(destination)  # декомпозиция кортежей
        print(non_priority, priority)
        if declared_value in ("нет", "", 0, "0"):
            fiz = non_priority[0] + round_as_excel(non_priority[1] * weight(item_weight, declared_value))
            yur = fiz
            for_declared_fiz = ''
            for_declared_yur = ''
            sep1 = ''
            sep2 = '/'
        else:
            fiz = 5.85 + non_priority[0] + round_as_excel(non_priority[1] * weight(item_weight, declared_value))
            print(non_priority[1] * weight(item_weight, declared_value))
            print(round_as_excel(non_priority[1] * weight(item_weight, declared_value)))
            yur = fiz
            for_declared_fiz = cost_for_declared_value(declared_value)
            fiz += cost_for_declared_value(declared_value)
            yur += cost_for_declared_value(declared_value)
            for_declared_yur = round_as_excel(cost_for_declared_value(declared_value)) * 1.2
            sep1, sep2 = "/", "/"

        item_vat_yur = round_as_excel(vat(yur))
        yur = round_as_excel(yur + item_vat_yur)
        rate = {
            'fiz': fiz,
            'yur': yur,
            'item_vat_yur': item_vat_yur,
            'for_declared_fiz': for_declared_fiz,
            'for_declared_yur': for_declared_yur,
            'rub': " руб.",
            'tracking': "да",
            'sep1': sep1,
            'sep2': sep2,
        }
        for key in rate:
            rate[key] = formatted(rate[key])
        price_row.append(rate)

    else:
        fiz = "Макс. вес 50 кг"
        sep1, sep2 = '', ''
        price_row.append({'fiz': fiz, 'sep1': sep1, 'sep2': sep2})
    return price_row
