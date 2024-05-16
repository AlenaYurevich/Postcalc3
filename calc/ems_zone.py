import os
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files/ems_zone.xlsx')
wb = load_workbook(filename=file_path)
sheet = wb.active


def find_ems_zone(point1, point2):
    x = get_column_letter(int(point1[1:]) + 1)  # get string from point1 without P
    y = str(int(point2[1:]) + 2)  # get row number from point2 without P
    return sheet[str(x + y)].value
